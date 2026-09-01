"""
app/routes_admin.py
====================
Sisi ADMIN KEJATI (v2 - disederhanakan):

  - Login sederhana
  - Periode Magang: dibuat sekaligus dengan upload data peserta dalam
    SATU form (nama periode, tanggal mulai/selesai via date picker,
    pilih template sertifikat, unggah Excel) - lihat periode_baru().
  - Data Peserta: tabel ringkas (No.Ref, Nama, NIM, Akun/No.WA, Waktu
    Diambil, Status, Aksi) langsung di halaman detail periode, lengkap
    dengan ringkasan status di bagian atas (tanpa halaman "Ringkasan"
    terpisah).
  - Template Sertifikat: unggah desain baru (PDF/gambar resolusi
    tinggi), kalibrasi posisi tiap field lewat form angka + pratinjau
    langsung (live preview), pilih template mana yang aktif dipakai per
    periode.

  Halaman "Log Aktivitas" sengaja TIDAK dibuat sebagai menu terpisah -
  status peserta (sudah/belum diambil, jumlah percobaan gagal) sudah
  cukup terlihat langsung di tabel Data Peserta.
"""
import os
import io
from datetime import datetime

from flask import (
    Blueprint, render_template, request, redirect, url_for, session,
    flash, current_app, send_file, jsonify, abort
)
from werkzeug.utils import secure_filename

from app.extensions import db
from app.models import Periode, Peserta, TemplateSertifikat
from app.auth import admin_required
from app.utils import baca_excel_peserta, proses_upload_peserta, cek_kode_verifikasi
from app.certgen.generator import (
    generate_certificate_image, DEFAULT_FIELD_CONFIG,
    FONT_NAME_DEFAULT, FONT_BOLD_DEFAULT, FONT_REG_DEFAULT,
)

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")


# ---------------------------------------------------------------- login --
@admin_bp.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")
        if (username == current_app.config["ADMIN_USERNAME"]
                and password == current_app.config["ADMIN_PASSWORD"]):
            session["is_admin"] = True
            session["admin_username"] = username
            flash("Berhasil masuk.", "success")
            tujuan = request.args.get("next") or url_for("admin.periode_list")
            return redirect(tujuan)
        flash("Username atau password salah.", "error")
    return render_template("admin/login.html")


@admin_bp.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("admin.login"))


# --------------------------------------------------------- periode list --
@admin_bp.route("/")
@admin_required
def index():
    return redirect(url_for("admin.periode_list"))


@admin_bp.route("/periode")
@admin_required
def periode_list():
    periode_list_ = Periode.query.order_by(Periode.dibuat_at.desc()).all()
    ringkasan = []
    for pr in periode_list_:
        total = len(pr.peserta)
        sudah = sum(1 for x in pr.peserta if x.status == "terkirim")
        ringkasan.append({"periode": pr, "total": total, "sudah": sudah, "belum": total - sudah})
    return render_template("admin/periode_list.html", ringkasan=ringkasan)


@admin_bp.route("/periode/baru", methods=["GET", "POST"])
@admin_required
def periode_baru():
    if request.method == "GET":
        template_list = TemplateSertifikat.query.filter_by(aktif=True).order_by(TemplateSertifikat.nama_template).all()
        return render_template("admin/periode_form.html", template_list=template_list, periode=None)

    nama = request.form.get("nama_periode", "").strip()
    tgl_mulai = request.form.get("tanggal_mulai") or None
    tgl_selesai = request.form.get("tanggal_selesai") or None
    template_id = request.form.get("template_id") or None

    if not nama:
        flash("Nama periode tidak boleh kosong.", "error")
        return redirect(url_for("admin.periode_baru"))

    periode = Periode(
        nama_periode=nama,
        tanggal_mulai=datetime.strptime(tgl_mulai, "%Y-%m-%d").date() if tgl_mulai else None,
        tanggal_selesai=datetime.strptime(tgl_selesai, "%Y-%m-%d").date() if tgl_selesai else None,
        template_id=int(template_id) if template_id else None,
        aktif=True,
    )
    db.session.add(periode)

    file_excel = request.files.get("file_excel")
    if file_excel and file_excel.filename:
        # PENTING: periode belum di-commit sampai titik ini. flush() dipakai
        # supaya periode.id tersedia (dibutuhkan relasi peserta) TANPA
        # menyimpannya permanen. Kalau proses upload Excel gagal di bawah -
        # baik karena file-nya sendiri bermasalah maupun error sistem lain -
        # seluruh transaksi (termasuk periode yang baru dibuat) di-rollback,
        # supaya tidak ada lagi "periode kosong" yang menumpuk tiap kali
        # admin mencoba ulang upload yang gagal (lihat riwayat perbaikan bug
        # No.Ref duplikat).
        try:
            db.session.flush()
            data_rows = baca_excel_peserta(file_excel)
            ditambahkan, dilewati = proses_upload_peserta(periode, data_rows, commit=False)
            db.session.commit()
            pesan = f"Periode '{nama}' dibuat dengan {ditambahkan} peserta."
            if dilewati:
                pesan += f" ({dilewati} baris dilewati karena NIM sudah ada)"
            flash(pesan, "success")
        except ValueError as e:
            db.session.rollback()
            flash(f"Gagal membuat periode: {e} Periode belum tersimpan - silakan perbaiki file Excel-nya dan coba lagi.", "error")
            return redirect(url_for("admin.periode_baru"))
        except Exception as e:
            db.session.rollback()
            flash(f"Gagal membuat periode karena kesalahan sistem ({e}). Periode belum tersimpan - silakan coba lagi.", "error")
            return redirect(url_for("admin.periode_baru"))
    else:
        db.session.commit()
        flash(f"Periode '{nama}' dibuat. Belum ada data peserta - unggah lewat halaman detail periode.", "success")

    return redirect(url_for("admin.periode_detail", periode_id=periode.id))


# ------------------------------------------------------- detail periode --
@admin_bp.route("/periode/<int:periode_id>")
@admin_required
def periode_detail(periode_id):
    periode = Periode.query.get_or_404(periode_id)
    peserta_list = Peserta.query.filter_by(periode_id=periode_id).order_by(Peserta.no_ref).all()

    total = len(peserta_list)
    sudah = sum(1 for p in peserta_list if p.status == "terkirim")
    belum = total - sudah
    gagal = sum(p.percobaan_gagal for p in peserta_list)

    return render_template(
        "admin/periode_detail.html",
        periode=periode, peserta_list=peserta_list,
        total=total, sudah=sudah, belum=belum, gagal=gagal,
        template_list=TemplateSertifikat.query.filter_by(aktif=True).order_by(TemplateSertifikat.nama_template).all(),
    )


@admin_bp.route("/periode/<int:periode_id>/upload-tambahan", methods=["POST"])
@admin_required
def upload_tambahan(periode_id):
    periode = Periode.query.get_or_404(periode_id)
    file_excel = request.files.get("file_excel")
    if not file_excel or not file_excel.filename:
        flash("Pilih file Excel terlebih dahulu.", "error")
        return redirect(url_for("admin.periode_detail", periode_id=periode_id))

    try:
        data_rows = baca_excel_peserta(file_excel)
        ditambahkan, dilewati = proses_upload_peserta(periode, data_rows)
        pesan = f"{ditambahkan} peserta baru ditambahkan."
        if dilewati:
            pesan += f" ({dilewati} baris dilewati karena NIM sudah ada di periode ini)"
        flash(pesan, "success")
    except ValueError as e:
        db.session.rollback()
        flash(f"Upload gagal: {e}", "error")
    except Exception as e:
        db.session.rollback()
        flash(f"Upload gagal karena kesalahan sistem ({e}). Silakan coba lagi.", "error")

    return redirect(url_for("admin.periode_detail", periode_id=periode_id))


@admin_bp.route("/periode/<int:periode_id>/edit", methods=["POST"])
@admin_required
def periode_edit(periode_id):
    """Panel 'Pengaturan Periode' berisi Nama Periode, Tanggal Terbit
    Sertifikat, & Template Sertifikat - tiga field yang benar-benar
    berdampak:
      - Nama: supaya typo saat membuat periode tidak harus hapus-buat-ulang.
      - Tanggal Terbit (kolom tanggal_selesai): DICETAK LANGSUNG di
        sertifikat lewat field dinamis "tanggal" (lihat certgen/
        generator.py & format_tanggal_indonesia di utils.py) - jadi
        field ini sempat dihapus dari panel saat masih dianggap murni
        kosmetik, sekarang dikembalikan karena sudah punya fungsi nyata.
      - Template: supaya admin bisa ganti desain untuk sisa peserta yang
        belum mengambil, tanpa mengubah PDF yang sudah terkirim.
    Tanggal Mulai TIDAK dikembalikan ke panel ini (tetap tersimpan &
    tampil di kartu periode, cuma memang tidak dipakai logika apa pun
    sampai saat ini) - kalau nanti ternyata dibutuhkan juga, tinggal
    ditambahkan dengan pola yang sama."""
    periode = Periode.query.get_or_404(periode_id)
    periode.nama_periode = request.form.get("nama_periode", periode.nama_periode).strip()
    tgl_terbit = request.form.get("tanggal_terbit") or None
    periode.tanggal_selesai = datetime.strptime(tgl_terbit, "%Y-%m-%d").date() if tgl_terbit else None
    template_id = request.form.get("template_id") or None
    periode.template_id = int(template_id) if template_id else None
    db.session.commit()
    flash("Pengaturan periode diperbarui.", "success")
    return redirect(url_for("admin.periode_detail", periode_id=periode_id))


@admin_bp.route("/periode/<int:periode_id>/hapus", methods=["POST"])
@admin_required
def periode_hapus(periode_id):
    """Hapus periode beserta seluruh data peserta di dalamnya (cascade -
    lihat relationship Periode.peserta di models.py). Dipakai terutama
    untuk membersihkan periode kosong/gagal (mis. sisa dari upload Excel
    yang dulu gagal sebelum bug penomoran No.Ref diperbaiki)."""
    periode = Periode.query.get_or_404(periode_id)
    nama = periode.nama_periode
    jumlah_peserta = len(periode.peserta)
    db.session.delete(periode)
    db.session.commit()
    flash(f"Periode '{nama}' beserta {jumlah_peserta} data peserta di dalamnya sudah dihapus.", "success")
    return redirect(url_for("admin.periode_list"))


@admin_bp.route("/cek-sertifikat")
@admin_required
def cek_sertifikat_cepat():
    """Endpoint JSON untuk quick-action 'Cek Keaslian Sertifikat' yang
    muncul sebagai pop-up di dashboard admin (lihat base_admin.html).
    Memakai logika yang sama persis dengan halaman publik /cek-sertifikat
    lewat helper cek_kode_verifikasi() supaya hasilnya selalu konsisten."""
    kode = request.args.get("kode", "").strip()
    hasil = cek_kode_verifikasi(kode)
    return jsonify(hasil or {"valid": False})


@admin_bp.route("/peserta/<int:peserta_id>/reset", methods=["POST"])
@admin_required
def reset_peserta(peserta_id):
    peserta = Peserta.query.get_or_404(peserta_id)
    peserta.status = "belum_diambil"
    peserta.waktu_diambil = None
    peserta.percobaan_gagal = 0
    db.session.commit()
    flash(f"Status '{peserta.nama}' direset ke Belum Diambil.", "success")
    return redirect(url_for("admin.periode_detail", periode_id=peserta.periode_id))


@admin_bp.route("/peserta/<int:peserta_id>/hapus", methods=["POST"])
@admin_required
def hapus_peserta(peserta_id):
    peserta = Peserta.query.get_or_404(peserta_id)
    periode_id = peserta.periode_id
    db.session.delete(peserta)
    db.session.commit()
    flash("Data peserta dihapus.", "success")
    return redirect(url_for("admin.periode_detail", periode_id=periode_id))


@admin_bp.route("/contoh-excel")
@admin_required
def unduh_contoh_excel():
    """Unduh file Excel contoh dengan kolom yang benar & tampilan rapi
    (lebar kolom pas, header berwarna, border jelas) supaya admin tidak
    perlu menebak-nebak format, dan hasilnya enak dibaca - bukan sekadar
    data mentah berdempetan seperti file CSV."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Data Peserta"

    kolom = ["Nama", "NIM", "Fakultas", "Universitas", "Email", "No. WA"]
    contoh_data = [
        ["Muhammad Naufal Nashir", "2305110041", "Ilmu Komputer", "Universitas Diponegoro",
         "naufal.nashir@example.com", "081234567890"],
        ["Clara Angelina Susanto Wijaya", "2305110048", "Ilmu Sosial dan Ilmu Politik",
         "Universitas Negeri Semarang", "clara.angelina@example.com", "081234567891"],
        ["Siti Aisyah Ramadhani", "2305110042", "Hukum", "Universitas Diponegoro",
         "siti.aisyah@example.com", "081234567892"],
    ]

    # --- Judul & petunjuk singkat di atas tabel ---
    ws.merge_cells("A1:F1")
    ws["A1"] = "Contoh Data Peserta Magang — Kejaksaan Tinggi Jawa Tengah"
    ws["A1"].font = Font(bold=True, size=13, color="2E1440")
    ws.merge_cells("A2:F2")
    ws["A2"] = "Hapus baris contoh di bawah, lalu isi dengan data peserta yang sebenarnya. Jangan mengubah nama kolom di baris 4."
    ws["A2"].font = Font(italic=True, size=9, color="6B7280")
    ws.row_dimensions[3].height = 6  # spasi kecil sebelum header tabel

    # --- Header tabel (baris 4) ---
    header_row = 4
    header_fill = PatternFill(start_color="5B2C82", end_color="5B2C82", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, judul in enumerate(kolom, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=judul)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # --- Baris data contoh ---
    for r, baris in enumerate(contoh_data, start=header_row + 1):
        for c, nilai in enumerate(baris, start=1):
            cell = ws.cell(row=r, column=c, value=nilai)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if r % 2 == 0:
                cell.fill = PatternFill(start_color="F5F3FB", end_color="F5F3FB", fill_type="solid")

    # --- Lebar kolom otomatis mengikuti isi terpanjang (tidak berdempetan) ---
    lebar_minimum = [26, 14, 26, 30, 28, 16]
    for col_idx, judul in enumerate(kolom, start=1):
        panjang_maks = max(
            [len(judul)] + [len(str(baris[col_idx - 1])) for baris in contoh_data]
        )
        lebar = max(panjang_maks + 4, lebar_minimum[col_idx - 1])
        ws.column_dimensions[get_column_letter(col_idx)].width = lebar

    ws.row_dimensions[header_row].height = 22
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)  # header tetap terlihat saat scroll

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="Contoh_Data_Peserta.xlsx",
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ============================================================ TEMPLATE ==
@admin_bp.route("/template")
@admin_required
def template_list():
    templates = TemplateSertifikat.query.order_by(TemplateSertifikat.dibuat_at.desc()).all()
    return render_template("admin/template_list.html", templates=templates)


@admin_bp.route("/template/<int:template_id>/gambar")
@admin_required
def template_preview_image(template_id):
    from flask import send_file as _send_file
    tpl = TemplateSertifikat.query.get_or_404(template_id)
    return _send_file(tpl.preview_path)


@admin_bp.route("/template/<int:template_id>/font/<jenis>")
@admin_required
def template_font_file(template_id, jenis):
    """Sajikan file font (kustom milik template, atau bawaan sistem kalau
    template belum unggah yang kustom) ke browser - dipakai halaman
    kalibrasi untuk memuat font ASLI lewat @font-face, supaya pratinjau
    langsung di kanvas (drag & ubah ukuran) tampil dengan font yang
    benar-benar sama seperti hasil render PDF sebenarnya, bukan cuma
    font browser sembarangan."""
    from flask import send_file as _send_file
    tpl = TemplateSertifikat.query.get_or_404(template_id)
    peta = {
        "nama": tpl.font_nama_path or FONT_NAME_DEFAULT,
        "bold": tpl.font_bold_path or FONT_BOLD_DEFAULT,
        "reg": tpl.font_reg_path or FONT_REG_DEFAULT,
        "tanggal": tpl.font_tanggal_path or FONT_REG_DEFAULT,
    }
    path = peta.get(jenis)
    if not path or not os.path.exists(path):
        abort(404)
    return _send_file(path)


ALLOWED_TEMPLATE_EXT = {"pdf", "png", "jpg", "jpeg"}
ALLOWED_FONT_EXT = {"ttf", "otf"}


def _ext_ok(filename, allowed):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in allowed


@admin_bp.route("/template/baru", methods=["GET", "POST"])
@admin_required
def template_baru():
    if request.method == "GET":
        return render_template("admin/template_form.html")

    nama = request.form.get("nama_template", "").strip()
    file_desain = request.files.get("file_desain")

    if not nama or not file_desain or not file_desain.filename:
        flash("Nama template dan file desain wajib diisi.", "error")
        return redirect(url_for("admin.template_baru"))

    if not _ext_ok(file_desain.filename, ALLOWED_TEMPLATE_EXT):
        flash("Format file desain harus PDF, PNG, atau JPG.", "error")
        return redirect(url_for("admin.template_baru"))

    os.makedirs(current_app.config["UPLOAD_TEMPLATE_DIR"], exist_ok=True)
    fname = secure_filename(file_desain.filename)
    stamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    raw_path = os.path.join(current_app.config["UPLOAD_TEMPLATE_DIR"], f"{stamp}_{fname}")
    file_desain.save(raw_path)

    # Kalau PDF, render ke PNG resolusi tinggi (300 DPI) dulu.
    if raw_path.lower().endswith(".pdf"):
        try:
            from pdf2image import convert_from_path
            pages = convert_from_path(raw_path, dpi=300)
            preview_path = raw_path.rsplit(".", 1)[0] + ".png"
            pages[0].save(preview_path, "PNG")
        except Exception as e:
            flash(
                f"Gagal mengonversi PDF ke gambar ({e}). Pastikan Poppler sudah "
                f"terpasang (lihat README bagian Prasyarat), atau unggah file "
                f"PNG/JPG langsung sebagai alternatif.", "error"
            )
            return redirect(url_for("admin.template_baru"))
    else:
        preview_path = raw_path

    # Peringatan resolusi rendah, supaya hasil cetak tetap tajam (HD).
    from PIL import Image
    with Image.open(preview_path) as im:
        w, h = im.size
    if w < 2400:
        flash(
            f"Perhatian: lebar gambar hanya {w}px. Untuk hasil sertifikat yang "
            f"tajam saat dicetak, disarankan unggah desain dengan lebar minimal "
            f"~2400px (setara 300 DPI ukuran A4). Template tetap tersimpan dan "
            f"bisa dipakai.", "error"
        )

    # Font kustom (opsional). Form sekarang cuma minta 3 file (Nama,
    # Identitas, Tanggal) - "Identitas" dipakai untuk font_bold_path
    # (label mis. "NIM :") maupun font_reg_path (isian mis. "2305110041")
    # SEKALIGUS, karena di desain resmi Kejati Jateng keduanya memang
    # font yang sama; menyimpannya di 2 kolom terpisah tetap dipertahankan
    # di database supaya template lama yang sudah terlanjur pakai font
    # label & nilai BERBEDA tidak rusak (tidak ada field di form baru
    # untuk mengatur itu lagi, tapi datanya tidak dihapus).
    os.makedirs(current_app.config["UPLOAD_FONT_DIR"], exist_ok=True)
    font_paths = {}
    for form_key, target_fields in [
        ("font_nama", ["font_nama_path"]),
        ("font_identitas", ["font_bold_path", "font_reg_path"]),
        ("font_tanggal", ["font_tanggal_path"]),
    ]:
        f = request.files.get(form_key)
        if f and f.filename:
            if not _ext_ok(f.filename, ALLOWED_FONT_EXT):
                flash(f"Font harus berformat .ttf atau .otf (dilewati: {f.filename}).", "error")
                continue
            fn = secure_filename(f.filename)
            fpath = os.path.join(current_app.config["UPLOAD_FONT_DIR"], f"{stamp}_{fn}")
            f.save(fpath)
            for target in target_fields:
                font_paths[target] = fpath

    tpl = TemplateSertifikat(
        nama_template=nama,
        preview_path=preview_path,
        font_nama_path=font_paths.get("font_nama_path"),
        font_bold_path=font_paths.get("font_bold_path"),
        font_reg_path=font_paths.get("font_reg_path"),
        font_tanggal_path=font_paths.get("font_tanggal_path"),
        aktif=True,
    )
    tpl.set_field_config(DEFAULT_FIELD_CONFIG)
    db.session.add(tpl)
    db.session.commit()

    # --- Deteksi otomatis posisi field dari teks placeholder (kalau ada) ---
    # Best-effort: field dengan label teks biasa (NIM/Fakultas/Universitas)
    # biasanya terdeteksi andal; field bergaya kaligrafi (Nama) kadang tidak.
    # Field yang tidak terdeteksi tetap memakai nilai bawaan dan bisa
    # diatur manual di halaman kalibrasi seperti biasa.
    try:
        from app.certgen.ocr_detect import deteksi_dan_bersihkan
        config_baru, laporan = deteksi_dan_bersihkan(tpl.preview_path, tpl.get_field_config())
        tpl.set_field_config(config_baru)
        db.session.commit()

        terdeteksi = [k for k, v in laporan.items() if v]
        tidak_terdeteksi = [k for k, v in laporan.items() if not v]
        if terdeteksi:
            pesan = f"Terdeteksi otomatis: {', '.join(terdeteksi)}."
            if tidak_terdeteksi:
                pesan += f" Perlu diatur manual: {', '.join(tidak_terdeteksi)}."
            flash(f"Template '{nama}' berhasil diunggah. {pesan}", "success")
        else:
            flash(
                f"Template '{nama}' berhasil diunggah. Tidak ada label placeholder "
                f"(Nama/NIM/Fakultas/Universitas) yang terdeteksi otomatis - silakan "
                f"atur posisi secara manual di halaman kalibrasi.", "success"
            )
    except Exception as e:
        flash(
            f"Template '{nama}' berhasil diunggah, tapi deteksi otomatis gagal dijalankan "
            f"({e}). Silakan atur posisi secara manual.", "error"
        )

    return redirect(url_for("admin.template_kalibrasi", template_id=tpl.id))


@admin_bp.route("/template/<int:template_id>/bersihkan-area", methods=["POST"])
@admin_required
def template_bersihkan_area(template_id):
    """Dipanggil tombol 'Bersihkan Area Ini' per field di halaman
    kalibrasi - untuk field yang tidak (atau salah) terdeteksi otomatis,
    supaya area teks lama tetap bisa dibersihkan tanpa perlu mengedit
    ulang file desain di luar sistem."""
    tpl = TemplateSertifikat.query.get_or_404(template_id)
    data = request.get_json(force=True)
    try:
        from app.certgen.ocr_detect import bersihkan_area
        bersihkan_area(
            tpl.preview_path,
            x_frac=float(data["x"]), y_frac=float(data["y"]),
            lebar_frac=float(data["lebar"]), tinggi_frac=float(data["tinggi"]),
        )
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


@admin_bp.route("/template/<int:template_id>/deteksi-ulang", methods=["POST"])
@admin_required
def template_deteksi_ulang(template_id):
    tpl = TemplateSertifikat.query.get_or_404(template_id)
    try:
        from app.certgen.ocr_detect import deteksi_dan_bersihkan
        config_baru, laporan = deteksi_dan_bersihkan(tpl.preview_path, tpl.get_field_config())
        tpl.set_field_config(config_baru)
        db.session.commit()

        terdeteksi = [k for k, v in laporan.items() if v]
        tidak_terdeteksi = [k for k, v in laporan.items() if not v]
        if terdeteksi:
            pesan = f"Terdeteksi ulang: {', '.join(terdeteksi)}."
            if tidak_terdeteksi:
                pesan += f" Masih perlu manual: {', '.join(tidak_terdeteksi)}."
            flash(pesan, "success")
        else:
            flash("Tidak ada label placeholder yang terdeteksi. Silakan atur manual.", "error")
    except Exception as e:
        flash(f"Deteksi otomatis gagal dijalankan ({e}).", "error")
    return redirect(url_for("admin.template_kalibrasi", template_id=tpl.id))


def _path_temp_diff(template_id, jenis):
    """jenis: 'contoh' atau 'kosong'. Lokasi file sementara khusus fitur
    deteksi 2-gambar, satu set per template (ditimpa kalau admin unggah
    ulang) - dibersihkan/dipindah permanen saat 'Terapkan Hasil' diklik."""
    folder = os.path.join(current_app.config["UPLOAD_TEMPLATE_DIR"], "_diff_tmp")
    os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, f"{template_id}_{jenis}.png")


@admin_bp.route("/template/<int:template_id>/deteksi-diff", methods=["POST"])
@admin_required
def template_deteksi_diff(template_id):
    """Langkah 1 dari fitur 'Deteksi dari 2 Gambar': terima 2 file
    (versi contoh + versi polos), simpan sementara, jalankan diff_detect,
    kembalikan daftar kotak yang terdeteksi (JSON) untuk admin beri label
    field lewat UI di halaman kalibrasi. Belum mengubah apa pun secara
    permanen di template - itu baru terjadi kalau admin klik "Terapkan
    Hasil" (lihat template_deteksi_diff_terapkan)."""
    tpl = TemplateSertifikat.query.get_or_404(template_id)
    f_contoh = request.files.get("gambar_contoh")
    f_kosong = request.files.get("gambar_kosong")
    if not f_contoh or not f_contoh.filename or not f_kosong or not f_kosong.filename:
        return jsonify({"ok": False, "error": "Unggah kedua file: versi contoh dan versi polos."}), 400
    for f in (f_contoh, f_kosong):
        if not _ext_ok(f.filename, {"png", "jpg", "jpeg"}):
            return jsonify({"ok": False, "error": f"Format {f.filename} tidak didukung - pakai PNG atau JPG."}), 400

    path_contoh = _path_temp_diff(template_id, "contoh")
    path_kosong = _path_temp_diff(template_id, "kosong")
    f_contoh.save(path_contoh)
    f_kosong.save(path_kosong)

    try:
        from app.certgen.diff_detect import deteksi_dari_dua_gambar
        kotak, diresize = deteksi_dari_dua_gambar(path_contoh, path_kosong)
    except Exception as e:
        return jsonify({"ok": False, "error": f"Deteksi gagal: {e}"}), 400

    if not kotak:
        return jsonify({"ok": False, "error": "Tidak ada perbedaan terdeteksi di antara kedua gambar. Pastikan versi contoh & polos memang berbeda teksnya."}), 400

    return jsonify({
        "ok": True, "kotak": kotak, "diresize": diresize,
        "gambar_contoh_url": url_for("admin.template_deteksi_diff_temp", template_id=template_id, jenis="contoh"),
    })


@admin_bp.route("/template/<int:template_id>/deteksi-diff/temp/<jenis>")
@admin_required
def template_deteksi_diff_temp(template_id, jenis):
    """Sajikan file sementara (contoh/kosong) untuk ditampilkan di modal
    pratinjau kotak deteksi - dipanggil browser lewat <img src=...>."""
    from flask import send_file as _send_file
    path = _path_temp_diff(template_id, jenis if jenis in ("contoh", "kosong") else "contoh")
    if not os.path.exists(path):
        abort(404)
    return _send_file(path)


@admin_bp.route("/template/<int:template_id>/deteksi-diff/terapkan", methods=["POST"])
@admin_required
def template_deteksi_diff_terapkan(template_id):
    """Langkah 2: admin sudah memberi label field ke tiap kotak lewat
    dropdown di UI. Terima pemetaan {field: [id_kotak, ...]}, lalu:
      1. Gabungkan bounding box semua kotak yang dilabeli field yang sama
         (kalau OCR/diff memecah satu field jadi >1 kotak, mis. label
         "NIM :" & nilainya kepisah) jadi satu titik tengah + perkiraan
         ukuran font dari tinggi kotak.
      2. Update field_config template dengan posisi & ukuran baru itu.
      3. Ganti preview_path template ke gambar VERSI POLOS yang barusan
         diunggah (karena sudah terbukti bersih - hasil diff yang jadi
         dasar deteksi ini) - jadi tidak perlu lagi "Bersihkan Area"
         manual, sekaligus posisi & background beres dalam satu langkah.
    """
    tpl = TemplateSertifikat.query.get_or_404(template_id)
    data = request.get_json(force=True)
    pemetaan = data.get("pemetaan", {})   # {"nama": [0], "nim": [1], ...}
    kotak_list = data.get("kotak", [])    # daftar kotak asli (dikirim balik dari klien)

    path_kosong = _path_temp_diff(template_id, "kosong")
    if not os.path.exists(path_kosong):
        return jsonify({"ok": False, "error": "File sementara sudah tidak ada, silakan unggah ulang kedua gambar."}), 400

    kotak_by_id = {k["id"]: k for k in kotak_list}
    config = tpl.get_field_config()
    # Template lama (dibuat sebelum field seperti "tanggal" ditambahkan ke
    # sistem) belum tentu punya semua key di field_config_json tersimpannya.
    # Isi dulu dengan bawaan supaya field APAPUN yang admin pilih di
    # dropdown pelabelan tetap bisa diproses, bukan malah dilewati diam-diam.
    for k, v in DEFAULT_FIELD_CONFIG.items():
        config.setdefault(k, v)
    field_terupdate = []

    for field, id_list in pemetaan.items():
        if field not in config or not id_list:
            continue
        kotak_terpilih = [kotak_by_id[i] for i in id_list if i in kotak_by_id]
        if not kotak_terpilih:
            continue

        # Gabung bounding box (union) dari semua kotak yang dilabeli field ini
        x0 = min(k["x"] - k["w"] / 2 for k in kotak_terpilih)
        x1 = max(k["x"] + k["w"] / 2 for k in kotak_terpilih)
        y0 = min(k["y"] - k["h"] / 2 for k in kotak_terpilih)
        y1 = max(k["y"] + k["h"] / 2 for k in kotak_terpilih)
        x_tengah = round((x0 + x1) / 2, 4)
        y_tengah = round((y0 + y1) / 2, 4)
        tinggi = y1 - y0

        # Perkiraan ukuran font dari tinggi kotak (faktor 0.8 - tinggi
        # kotak biasanya sedikit lebih besar dari tinggi huruf kapital
        # karena ikut menangkap ascender/descender). Titik awal yang
        # jauh lebih baik daripada nilai bawaan sistem - admin masih
        # bebas menyesuaikan lewat stepper px seperti biasa.
        perkiraan_size = round(tinggi * 0.8, 4)

        if field == "nama":
            config["nama"]["x"] = x_tengah
            config["nama"]["y"] = y_tengah
            if perkiraan_size > 0:
                config["nama"]["max_size"] = perkiraan_size
                config["nama"]["min_size"] = round(perkiraan_size * 0.55, 4)
        elif field == "nomor":
            config["nomor"]["x"] = round(x1, 4)  # rata kanan -> pakai tepi kanan kotak
            config["nomor"]["y1"] = y_tengah
            config["nomor"]["y2"] = round(y_tengah + max(tinggi, 0.02), 4)
            config["nomor"]["tampilkan"] = True
            if perkiraan_size > 0:
                config["nomor"]["size"] = perkiraan_size
        else:
            config[field]["x"] = x_tengah
            config[field]["y"] = y_tengah
            if field == "tanggal":
                config["tanggal"]["tampilkan"] = True
            if perkiraan_size > 0 and "size" in config[field]:
                config[field]["size"] = perkiraan_size

        field_terupdate.append(field)

    tpl.set_field_config(config)

    # Pindahkan gambar "polos" jadi preview_path permanen template ini
    import shutil
    stamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    path_permanen = os.path.join(current_app.config["UPLOAD_TEMPLATE_DIR"], f"{stamp}_polos_{template_id}.png")
    shutil.copy(path_kosong, path_permanen)
    tpl.preview_path = path_permanen

    db.session.commit()

    return jsonify({"ok": True, "field_terupdate": field_terupdate, "config": config})


@admin_bp.route("/template/<int:template_id>/kalibrasi", methods=["GET"])
@admin_required
def template_kalibrasi(template_id):
    tpl = TemplateSertifikat.query.get_or_404(template_id)
    config = tpl.get_field_config()
    # Template yang dibuat sebelum fitur "Tanggal Terbit" ada belum punya
    # key "tanggal" di field_config_json tersimpannya - isi dengan nilai
    # bawaan supaya form kalibrasi tetap tampil normal (bisa langsung
    # disesuaikan posisinya), bukan kosong/error.
    config.setdefault("tanggal", DEFAULT_FIELD_CONFIG["tanggal"])
    return render_template("admin/template_kalibrasi.html", tpl=tpl, config=config)


@admin_bp.route("/template/<int:template_id>/kalibrasi/simpan", methods=["POST"])
@admin_required
def template_kalibrasi_simpan(template_id):
    tpl = TemplateSertifikat.query.get_or_404(template_id)
    data = request.get_json(force=True)
    tpl.set_field_config(data)
    db.session.commit()
    return jsonify({"ok": True})


@admin_bp.route("/template/<int:template_id>/kalibrasi/preview", methods=["POST"])
@admin_required
def template_kalibrasi_preview(template_id):
    """Render pratinjau langsung (belum disimpan) dengan data contoh,
    dipakai oleh tombol 'Render Uji Coba' di halaman kalibrasi."""
    tpl = TemplateSertifikat.query.get_or_404(template_id)
    config = request.get_json(force=True)

    im = generate_certificate_image(
        preview_path=tpl.preview_path,
        field_config=config,
        nama="Contoh Nama Panjang Peserta", nim="0000000000",
        fakultas="Contoh Fakultas", universitas="Contoh Universitas Perguruan Tinggi",
        no_sertifikat="001/CONTOH/VIII/2026", kode_verifikasi="KT26-CONTOH",
        verify_url="http://127.0.0.1:5000/cek-sertifikat?kode=KT26-CONTOH",
        tanggal_terbit="30 Agustus 2026",
        font_nama_path=tpl.font_nama_path, font_bold_path=tpl.font_bold_path, font_reg_path=tpl.font_reg_path,
        font_tanggal_path=tpl.font_tanggal_path,
    )
    buf = io.BytesIO()
    im.save(buf, "JPEG", quality=85)
    buf.seek(0)
    return send_file(buf, mimetype="image/jpeg")


@admin_bp.route("/template/<int:template_id>/nonaktifkan", methods=["POST"])
@admin_required
def template_nonaktifkan(template_id):
    tpl = TemplateSertifikat.query.get_or_404(template_id)
    tpl.aktif = not tpl.aktif
    db.session.commit()
    flash(f"Template '{tpl.nama_template}' {'diaktifkan' if tpl.aktif else 'dinonaktifkan'}.", "success")
    return redirect(url_for("admin.template_list"))
