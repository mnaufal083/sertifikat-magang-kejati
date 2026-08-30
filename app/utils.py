"""
app/utils.py
============
Fungsi bantu: penomoran sertifikat, OTP via Email, pengiriman
sertifikat via email, dan pembacaan file Excel data peserta.

>>> ALUR VERIFIKASI (v3) <<<
Peserta memasukkan Nama + NIM + No. WhatsApp pada portal publik. Kalau
cocok dengan data yang diunggah admin, kode OTP dikirim ke EMAIL
peserta yang terdaftar (No. WA hanya dipakai untuk mencocokkan
identitas, BUKAN untuk mengirim apa pun - sebelumnya OTP dikirim lewat
WhatsApp, tapi diganti ke email karena WhatsApp API berbayar per
pesan). Begitu OTP benar, sertifikat PDF langsung dibuat dan DIKIRIM
SEBAGAI LAMPIRAN EMAIL ke alamat yang sama (bukan diunduh langsung dari
browser) - meniru pola pengiriman sertifikat bootcamp/seminar resmi,
sekaligus memberi jejak pengiriman yang formal.

>>> SOLUSI NOMOR SERTIFIKAT <<<
Nomor referensi (No. Ref) & kode verifikasi dibuat SEKALI saat admin
mengunggah data peserta (lihat proses_upload_peserta di bawah), bukan
menunggu peserta mengklaim. Ini membuat "No. Ref" selalu terlihat di
tabel admin sejak data diunggah, dan nomor tersebut yang sama persis
dipakai sebagai "No. Sertifikat" begitu file PDF akhirnya dibuat.
"""
import random
import re
import secrets
import string
from datetime import datetime, timedelta

from app.extensions import db
from app.models import Peserta, OtpCode

ROMAWI_BULAN = ["", "I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII"]


# --------------------------------------------------------------- nomor --
def _nomor_urut_berikutnya():
    """Dihitung dari NOMOR TERTINGGI yang PERNAH dipakai di seluruh data
    peserta (bukan sekadar total baris yang tersisa saat ini), supaya
    penomoran tetap aman & tidak bertabrakan meskipun ada peserta yang
    sudah dihapus sebelumnya.

    (Riwayat bug: versi lama memakai Peserta.query.count(). Begitu ada
    peserta yang dihapus, hasil count() ikut mundur, sehingga nomor
    berikutnya yang dihitung ulang bisa jadi sama dengan nomor besar
    yang masih dipakai peserta lain yang TIDAK dihapus - menghasilkan
    IntegrityError "UNIQUE constraint failed: peserta.no_ref". Ambil
    MAX bukan COUNT supaya kasus ini tidak terjadi lagi.)

    Query hanya mengambil kolom no_ref (bukan seluruh baris) supaya
    tetap ringan meski jumlah peserta sudah banyak. db.session.query()
    di sini otomatis meng-autoflush peserta yang sudah ditambahkan (tapi
    belum di-commit) pada baris-baris sebelumnya di batch upload yang
    sama, sehingga penomoran tetap berurutan benar dalam satu kali
    proses upload Excel."""
    tertinggi = 0
    for (no_ref,) in db.session.query(Peserta.no_ref).all():
        m = re.match(r"^(\d+)/", no_ref or "")
        if m:
            tertinggi = max(tertinggi, int(m.group(1)))
    return tertinggi + 1


def buat_no_ref(tanggal=None):
    tanggal = tanggal or datetime.utcnow()
    urut = _nomor_urut_berikutnya()
    return f"{urut:03d}/PTJT.6/Mag.6/{ROMAWI_BULAN[tanggal.month]}/{tanggal.year}"


def buat_kode_verifikasi():
    tahun = str(datetime.utcnow().year)[-2:]
    acak = "".join(secrets.choice(string.ascii_uppercase + string.digits) for _ in range(6))
    return f"KT{tahun}-{acak}"


# --------------------------------------------------------------- waktu --
def ke_wib(dt):
    """Konversi datetime UTC (yang disimpan di database - praktik yang
    benar) ke WIB (UTC+7) HANYA untuk ditampilkan ke admin/peserta.
    Indonesia tidak menerapkan daylight saving, jadi offset +7 jam ini
    selalu tetap sepanjang tahun untuk Semarang/Jawa Tengah.

    Sebelumnya tampilan waktu (mis. 'Waktu Diambil' di tabel peserta)
    langsung mencetak datetime.utcnow() tanpa konversi, sehingga jamnya
    tampak mundur 7 jam dari waktu asli WIB saat admin melihatnya."""
    if dt is None:
        return None
    return dt + timedelta(hours=7)


_NAMA_BULAN_ID = [
    "", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember",
]


def format_tanggal_indonesia(tanggal):
    """Format objek date/datetime jadi teks tanggal Indonesia, mis.
    '30 Agustus 2026'. Dipakai untuk field Tanggal Terbit dinamis di
    sertifikat (lihat certgen/generator.py) - mengikuti tanggal_selesai
    periode magang, jadi tidak perlu ganti template tiap periode
    berbeda tanggalnya."""
    if tanggal is None:
        return ""
    return f"{tanggal.day} {_NAMA_BULAN_ID[tanggal.month]} {tanggal.year}"


# ---------------------------------------------------------- import data --
KOLOM_WAJIB = ["nama", "nim", "fakultas", "universitas", "email", "no_wa"]

# Nama kolom alternatif yang diterima di file Excel (tidak peka huruf besar/kecil)
ALIAS_KOLOM = {
    "nama": ["nama", "nama lengkap", "nama peserta"],
    "nim": ["nim", "nomor induk", "nim/nomor induk", "no induk"],
    "fakultas": ["fakultas"],
    "universitas": ["universitas", "kampus", "perguruan tinggi"],
    "email": ["email", "email terdaftar", "alamat email"],
    "no_wa": ["no_wa", "no wa", "nomor wa", "whatsapp", "no whatsapp", "no. wa", "nomor whatsapp"],
}


def _cocokkan_kolom(header_baris):
    """header_baris: list nama kolom mentah dari baris pertama Excel.
    Mengembalikan dict {field_internal: index_kolom}."""
    header_bersih = [str(h).strip().lower() if h else "" for h in header_baris]
    hasil = {}
    for field, alias_list in ALIAS_KOLOM.items():
        for idx, h in enumerate(header_bersih):
            if h in alias_list:
                hasil[field] = idx
                break
    return hasil


def baca_excel_peserta(file_stream):
    """Baca file .xlsx dan kembalikan list dict {nama, nim, fakultas,
    universitas, email, no_wa}. Melempar ValueError dengan pesan jelas
    kalau ada kolom wajib yang tidak ditemukan.

    Baris header (Nama/NIM/dst.) TIDAK harus persis di baris pertama -
    fungsi ini memindai beberapa baris pertama untuk menemukannya, supaya
    file yang punya judul/keterangan di atas tabel (seperti contoh yang
    diunduh dari "Unduh contoh file") tetap terbaca dengan benar."""
    from openpyxl import load_workbook

    wb = load_workbook(file_stream, read_only=True, data_only=True)
    ws = wb.active
    semua_baris = list(ws.iter_rows(values_only=True))

    if not semua_baris:
        raise ValueError("File Excel kosong.")

    BATAS_PINDAI = 10  # cukup untuk menampung judul/keterangan di atas tabel
    kolom_map = None
    baris_header_ke = None
    for idx, baris in enumerate(semua_baris[:BATAS_PINDAI]):
        percobaan = _cocokkan_kolom(baris)
        if all(f in percobaan for f in KOLOM_WAJIB):
            kolom_map = percobaan
            baris_header_ke = idx
            break

    if kolom_map is None:
        raise ValueError(
            "Baris header (Nama, NIM, Fakultas, Universitas, Email, No. WA) tidak "
            "ditemukan di " + str(BATAS_PINDAI) + " baris pertama file Excel. "
            "Pastikan nama kolom tsb ada persis sebagai satu baris di dekat "
            "bagian atas file, lalu coba unggah ulang."
        )

    baris_iter = iter(semua_baris[baris_header_ke + 1:])

    hasil = []
    for baris in baris_iter:
        if baris is None or all(v is None or str(v).strip() == "" for v in baris):
            continue  # lewati baris kosong
        row = {
            field: (str(baris[idx]).strip() if baris[idx] is not None else "")
            for field, idx in kolom_map.items()
        }
        if not row["nama"] or not row["nim"]:
            continue
        hasil.append(row)
    return hasil


def proses_upload_peserta(periode, data_rows, commit=True):
    """Simpan banyak baris peserta sekaligus ke satu periode, sekaligus
    membuatkan No. Ref & kode verifikasi untuk masing-masing. Melewati
    baris yang NIM-nya sudah ada di periode yang sama (mencegah duplikat
    kalau file diunggah dua kali).

    commit=False dipakai saat pemanggil (mis. periode_baru) ingin
    mengontrol sendiri kapan transaksi disimpan permanen - supaya kalau
    ada baris yang gagal di tengah proses, periode yang baru dibuat pun
    ikut batal (rollback) alih-alih tersimpan sebagai periode kosong.

    Mengembalikan (jumlah_ditambahkan, jumlah_dilewati)."""
    ditambahkan, dilewati = 0, 0
    nim_ada = {p.nim for p in periode.peserta}

    for row in data_rows:
        if row["nim"] in nim_ada:
            dilewati += 1
            continue
        p = Peserta(
            periode_id=periode.id,
            nama=row["nama"], nim=row["nim"],
            fakultas=row["fakultas"], universitas=row["universitas"],
            email=row["email"].lower(), no_wa=_normalisasi_no_wa(row["no_wa"]),
            no_ref=buat_no_ref(),
            kode_verifikasi=buat_kode_verifikasi(),
        )
        db.session.add(p)
        nim_ada.add(row["nim"])
        ditambahkan += 1

    if commit:
        db.session.commit()
    return ditambahkan, dilewati


def _normalisasi_no_wa(no_wa):
    """Rapikan format nomor WA: buang spasi/strip, ubah awalan 0 -> 62."""
    bersih = "".join(ch for ch in no_wa if ch.isdigit() or ch == "+")
    bersih = bersih.replace("+", "")
    if bersih.startswith("0"):
        bersih = "62" + bersih[1:]
    return bersih


# ------------------------------------------------------------------ OTP --
def cari_peserta_untuk_verifikasi(nama, nim, email):
    """Cari peserta yang cocok Nama + NIM + Email di SEMUA periode
    (peserta tidak perlu tahu/pilih periode-nya sendiri). Pencocokan nama
    tidak peka besar-kecil huruf & spasi berlebih; NIM harus identik;
    email dicocokkan tidak peka besar-kecil huruf (sesuai kebiasaan
    penulisan alamat email).

    Catatan: sebelumnya field pencocokan ketiga ini adalah No. WhatsApp,
    tapi diganti ke Email karena OTP & sertifikat memang selalu dikirim
    lewat email - jadi dari sudut pandang peserta, memasukkan email di
    awal lebih masuk akal & konsisten dibanding nomor WA yang sebenarnya
    tidak pernah dipakai mengirim apa pun."""
    nim_bersih = nim.strip()
    email_bersih = email.strip().lower()
    nama_bersih = " ".join(nama.strip().lower().split())

    kandidat = Peserta.query.filter_by(nim=nim_bersih).all()
    for p in kandidat:
        nama_p = " ".join(p.nama.strip().lower().split())
        if nama_p == nama_bersih and p.email.strip().lower() == email_bersih:
            return p
    return None


def buat_dan_kirim_otp(peserta, expiry_minutes=5, demo_mode=True):
    """Buat kode OTP 6 digit baru untuk peserta, simpan hash-nya, dan
    kirim ke EMAIL peserta yang terdaftar (atau tampilkan di layar bila
    demo_mode). No. WA peserta tetap dipakai untuk mencocokkan identitas
    di awal (lihat cari_peserta_untuk_verifikasi), tapi tidak lagi dipakai
    untuk mengirim apa pun."""
    kode = f"{random.randint(0, 999999):06d}"

    otp = OtpCode(
        peserta_id=peserta.id,
        kedaluwarsa_at=datetime.utcnow() + timedelta(minutes=expiry_minutes),
    )
    otp.set_kode(kode)
    if demo_mode:
        otp.kode_plain_demo = kode
    db.session.add(otp)
    db.session.commit()

    if demo_mode:
        print(f"[DEMO] Kode OTP untuk {peserta.email}: {kode} (berlaku {expiry_minutes} menit)")
    else:
        kirim_email_otp(peserta.email, peserta.nama, kode, expiry_minutes)

    return otp


def kirim_email_otp(email_tujuan, nama, kode, expiry_minutes):
    """Kirim kode OTP verifikasi ke email peserta lewat SMTP. Dipanggil
    saat EMAIL_DEMO_MODE=false (lihat kirim_sertifikat_email untuk fungsi
    pengiriman sertifikat, yang memakai SMTP yang sama)."""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.utils import formatdate, make_msgid
    from flask import current_app

    cfg = current_app.config
    if not cfg.get("SMTP_HOST"):
        raise RuntimeError(
            "SMTP belum dikonfigurasi. Isi SMTP_HOST/SMTP_USERNAME/SMTP_PASSWORD "
            "di .env, atau set EMAIL_DEMO_MODE=true untuk mode demo."
        )

    pesan = MIMEMultipart()
    pesan["Subject"] = "Kode Verifikasi OTP — Sertifikat Magang Kejati Jateng"
    pesan["From"] = cfg["SMTP_FROM"]
    pesan["To"] = email_tujuan
    # Header Date & Message-ID eksplisit - tanpa ini, email lebih rentan
    # ditandai spam oleh sebagian besar penyedia email (termasuk Gmail),
    # karena dianggap salah satu ciri email yang dikirim asal-asalan oleh bot.
    pesan["Date"] = formatdate(localtime=True)
    pesan["Message-ID"] = make_msgid(domain=cfg["SMTP_FROM"].split("@")[-1])

    isi = (
        f"Yth. {nama},\n\n"
        f"Berikut kode verifikasi (OTP) untuk mengambil sertifikat magang "
        f"Anda di lingkungan Kejaksaan Tinggi Jawa Tengah:\n\n"
        f"    {kode}\n\n"
        f"Kode ini berlaku selama {expiry_minutes} menit dan hanya bisa "
        f"dipakai satu kali. Jangan berikan kode ini kepada siapa pun.\n\n"
        f"Jika Anda tidak merasa meminta kode ini, abaikan saja email ini.\n\n"
        f"Hormat kami,\n"
        f"Bidang Pembinaan\n"
        f"Kejaksaan Tinggi Jawa Tengah\n\n"
        f"--\n"
        f"Email ini dikirim otomatis oleh sistem, mohon tidak membalas ke alamat ini."
    )
    pesan.attach(MIMEText(isi, "plain"))

    with smtplib.SMTP(cfg["SMTP_HOST"], cfg["SMTP_PORT"]) as server:
        server.starttls()
        try:
            server.login(cfg["SMTP_USERNAME"], cfg["SMTP_PASSWORD"])
        except smtplib.SMTPAuthenticationError:
            raise RuntimeError(
                "Login SMTP ditolak Gmail (kredensial salah). Kalau memakai Gmail, "
                "SMTP_PASSWORD di .env HARUS berupa App Password 16 karakter "
                "(bukan password akun Gmail biasa) - lihat README bagian 8 untuk "
                "cara membuatnya."
            )
        server.sendmail(cfg["SMTP_FROM"], [email_tujuan], pesan.as_string())


# ------------------------------------------------------- kirim sertifikat --
def kirim_sertifikat_email(peserta, pdf_bytes, demo_mode=True):
    """Kirim file PDF sertifikat sebagai lampiran email resmi ke alamat
    terdaftar. Mode demo: file disimpan ke data/generated/ supaya tetap
    bisa diperiksa hasilnya tanpa SMTP asli (lihat route demo di
    routes_public.py)."""
    import os
    from flask import current_app

    if demo_mode:
        out_dir = os.path.join(current_app.root_path, "..", "data", "generated")
        os.makedirs(out_dir, exist_ok=True)
        filename = f"{peserta.no_ref.replace('/', '_')}.pdf"
        with open(os.path.join(out_dir, filename), "wb") as f:
            f.write(pdf_bytes)
        print(f"[DEMO] Sertifikat untuk {peserta.email} disimpan di data/generated/{filename}")
        return filename

    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.application import MIMEApplication
    from email.utils import formatdate, make_msgid

    cfg = current_app.config
    if not cfg.get("SMTP_HOST"):
        raise RuntimeError(
            "SMTP belum dikonfigurasi. Isi SMTP_HOST/SMTP_USERNAME/SMTP_PASSWORD "
            "di .env, atau set EMAIL_DEMO_MODE=true untuk mode demo."
        )

    pesan = MIMEMultipart()
    pesan["Subject"] = "Sertifikat Magang Resmi — Kejaksaan Tinggi Jawa Tengah"
    # Header Date & Message-ID eksplisit - salah satu sinyal umum yang
    # dicek filter spam; tanpa ini email lebih rentan ditandai spam.
    pesan["Date"] = formatdate(localtime=True)
    pesan["Message-ID"] = make_msgid(domain=cfg["SMTP_FROM"].split("@")[-1])
    pesan["From"] = cfg["SMTP_FROM"]
    pesan["To"] = peserta.email

    isi = (
        f"Yth. {peserta.nama},\n\n"
        f"Selamat! Identitas Anda telah berhasil diverifikasi. Bersama email ini "
        f"kami lampirkan sertifikat resmi magang Anda di lingkungan Kejaksaan "
        f"Tinggi Jawa Tengah.\n\n"
        f"No. Sertifikat   : {peserta.no_ref}\n"
        f"Kode Verifikasi  : {peserta.kode_verifikasi}\n\n"
        f"Keaslian sertifikat ini dapat diverifikasi kapan saja melalui kode QR "
        f"pada sertifikat, atau melalui halaman resmi kami dengan memasukkan "
        f"kode verifikasi di atas.\n\n"
        f"Terima kasih atas kontribusi Anda selama masa magang.\n\n"
        f"Hormat kami,\n"
        f"Bidang Pembinaan\n"
        f"Kejaksaan Tinggi Jawa Tengah\n\n"
        f"--\n"
        f"Email ini dikirim otomatis oleh sistem, mohon tidak membalas ke alamat ini."
    )
    pesan.attach(MIMEText(isi, "plain"))

    lampiran = MIMEApplication(pdf_bytes, _subtype="pdf")
    lampiran.add_header("Content-Disposition", "attachment",
                         filename=f"Sertifikat_Magang_{peserta.nama.replace(' ', '_')}.pdf")
    pesan.attach(lampiran)

    with smtplib.SMTP(cfg["SMTP_HOST"], cfg["SMTP_PORT"]) as server:
        server.starttls()
        try:
            server.login(cfg["SMTP_USERNAME"], cfg["SMTP_PASSWORD"])
        except smtplib.SMTPAuthenticationError:
            raise RuntimeError(
                "Login SMTP ditolak Gmail (kredensial salah). Kalau memakai Gmail, "
                "SMTP_PASSWORD di .env HARUS berupa App Password 16 karakter "
                "(bukan password akun Gmail biasa) - lihat README bagian 8 untuk "
                "cara membuatnya."
            )
        server.sendmail(cfg["SMTP_FROM"], [peserta.email], pesan.as_string())
    return None


# ---------------------------------------------------- cek keaslian ------
def cek_kode_verifikasi(kode):
    """Cek satu kode verifikasi sertifikat. Dipakai bersama oleh halaman
    publik (/cek-sertifikat) maupun quick-action "Cek Keaslian Sertifikat"
    di dashboard admin, supaya logikanya tidak dobel dan selalu konsisten.

    Mengembalikan dict hasil (tanggal sudah diformat jadi string supaya
    aman dipakai langsung di template maupun di-serialize ke JSON)."""
    kode = (kode or "").strip()
    if not kode:
        return None
    peserta = Peserta.query.filter_by(kode_verifikasi=kode).first()
    if peserta and peserta.status == "terkirim":
        return {
            "valid": True,
            "nama": peserta.nama, "nim": peserta.nim,
            "universitas": peserta.universitas,
            "no_sertifikat": peserta.no_ref,
            "tanggal_terbit": peserta.waktu_diambil.strftime("%d %B %Y") if peserta.waktu_diambil else "-",
            "periode": peserta.periode.nama_periode,
        }
    return {"valid": False}
